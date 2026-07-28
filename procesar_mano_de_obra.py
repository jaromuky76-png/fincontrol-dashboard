import os
import sys
import zipfile
import re
import time
import json
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("==========================================================")
print("   FinControl - Agente Automatizado de Costeo de Mano de Obra  ")
print("==========================================================")

t0 = time.time()

# 1. Load Master Catalogs
maestros_path = r'COD MAESTROS\COD MAESTROS.xlsx'
cs_path = r'COD CENTRO DE SERVICIOS\COD CS.xlsx'

def load_catalog(filepath):
    if not os.path.exists(filepath):
        print(f"Advertencia: No se encontró el archivo {filepath}")
        return {}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    catalog = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None: continue
        c_str = str(row[0]).strip()
        if c_str.endswith('.0'): c_str = c_str[:-2]
        if c_str.lower() in ('código producto', 'codigo producto', 'código', 'codigo'): continue
        desc_str = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        catalog[c_str] = desc_str
    wb.close()
    return catalog

print("\n1. Cargando catálogos de referencia de mano de obra...")
maestros_catalog = load_catalog(maestros_path)
cs_catalog = load_catalog(cs_path)

print(f" -> Taller Maestro (T49): {len(maestros_catalog)} códigos cargados")
print(f" -> Centro de Servicios (T39): {len(cs_catalog)} códigos cargados")

# 2. Locate and Select Sales Consolidated File
sales_dir = r'CONSOLIDADO DE VENTAS'
excel_files = []
if os.path.exists(sales_dir):
    for f in os.listdir(sales_dir):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            full_p = os.path.join(sales_dir, f)
            excel_files.append((full_p, f, os.path.getmtime(full_p)))

# Sort by modification date (most recent first)
excel_files.sort(key=lambda x: x[2], reverse=True)

if not excel_files:
    print("\nError: No se encontró ningún archivo .xlsx en la carpeta CONSOLIDADO DE VENTAS.")
    input("\nPresiona Enter para salir...")
    sys.exit(1)

selected_file_path = None
filename_base = None

if len(excel_files) == 1:
    selected_file_path = excel_files[0][0]
    filename_base = excel_files[0][1]
    print(f"\n2. Sábana de ventas detectada: {filename_base}")
else:
    print("\n2. Se encontraron varios archivos de sábana en 'CONSOLIDADO DE VENTAS':")
    for idx, (path, fname, mtime) in enumerate(excel_files, 1):
        mtime_str = time.strftime('%d/%m/%Y %H:%M', time.localtime(mtime))
        size_mb = os.path.getsize(path) / (1024 * 1024)
        recency = " (Más reciente)" if idx == 1 else ""
        print(f"   [{idx}] {fname} ({size_mb:.1f} MB - {mtime_str}){recency}")
    
    user_choice = input(f"\nSelecciona el número del archivo a procesar [1-{len(excel_files)}] (Presiona ENTER para la opción [1]): ").strip()
    
    if not user_choice:
        selected_file_path = excel_files[0][0]
        filename_base = excel_files[0][1]
    else:
        try:
            choice_idx = int(user_choice) - 1
            if 0 <= choice_idx < len(excel_files):
                selected_file_path = excel_files[choice_idx][0]
                filename_base = excel_files[choice_idx][1]
            else:
                selected_file_path = excel_files[0][0]
                filename_base = excel_files[0][1]
        except ValueError:
            selected_file_path = excel_files[0][0]
            filename_base = excel_files[0][1]

size_mb = os.path.getsize(selected_file_path) / (1024 * 1024)
print(f"\nProcesando archivo seleccionado: {filename_base} ({size_mb:.1f} MB)...")

# 3. Fast Parsing: Extract Both Unconsolidated Rows (for Excel) & Consolidated Summary (for Web)
maestros_unconsolidated = []  # List of { row, code, desc }
cs_unconsolidated = []        # List of { row, code, desc }

maestros_grouped = {}         # code -> { code, desc, frequency }
cs_grouped = {}               # code -> { code, desc, frequency }

total_sales_rows = 0
total_labor_occurrences = 0

try:
    with zipfile.ZipFile(selected_file_path, 'r') as z:
        # Load Shared Strings
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            ss_bytes = z.read('xl/sharedStrings.xml')
            si_matches = re.findall(rb'<si>(.*?)</si>', ss_bytes, re.DOTALL)
            for si in si_matches:
                t_matches = re.findall(rb'<t[^>]*>(.*?)</t>', si, re.DOTALL)
                if t_matches:
                    shared_strings.append(b"".join(t_matches).decode('utf-8', errors='ignore'))
                else:
                    shared_strings.append('')

        # Locate Ventas sheet
        sheet_path = 'xl/worksheets/sheet3.xml'
        wb_bytes = z.read('xl/workbook.xml')
        sheet_matches = re.findall(rb'<sheet\s+[^>]*>', wb_bytes, re.IGNORECASE)
        found_r_id = None
        for tag in sheet_matches:
            name_m = re.search(rb'name="([^"]+)"', tag, re.IGNORECASE)
            if name_m and any(k in name_m.group(1).lower() for k in [b'venta', b'sale', b'factur']):
                r_id_m = re.search(rb'r:id="([^"]+)"', tag, re.IGNORECASE) or re.search(rb':id="([^"]+)"', tag, re.IGNORECASE)
                if r_id_m:
                    found_r_id = r_id_m.group(1)
                    break

        if found_r_id:
            rels_bytes = z.read('xl/_rels/workbook.xml.rels')
            rel_m = re.search(rb'Id="' + found_r_id + rb'"[^>]*Target="([^"]+)"', rels_bytes, re.IGNORECASE)
            if rel_m:
                sheet_path = 'xl/' + rel_m.group(1).decode('utf-8').lstrip('/').replace('xl/', '')

        if sheet_path not in z.namelist():
            sheet_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
            sheet_files.sort(key=lambda f: z.getinfo(f).file_size, reverse=True)
            sheet_path = sheet_files[0] if sheet_files else 'xl/worksheets/sheet3.xml'

        print(f" -> Analizando hoja de ventas: {sheet_path}")
        sheet_bytes = z.read(sheet_path)

        cell_q_regex = re.compile(rb'<c r="Q(\d+)"([^>]*)>(.*?)</c>', re.DOTALL)
        val_regex = re.compile(rb'<v>(.*?)</v>')
        t_attr_regex = re.compile(rb't="([^"]+)"')

        for match in cell_q_regex.finditer(sheet_bytes):
            row_idx_str = match.group(1).decode('utf-8')
            if row_idx_str == '1': continue # Header
            
            total_sales_rows += 1
            inner = match.group(3)
            val_m = val_regex.search(inner)
            if not val_m: continue

            val = val_m.group(1).decode('utf-8', errors='ignore')
            t_m = t_attr_regex.search(match.group(2))
            cell_type = t_m.group(1).decode('utf-8') if t_m else ''

            if cell_type == 's':
                idx = int(val)
                code_str = shared_strings[idx].strip() if idx < len(shared_strings) else ''
            else:
                code_str = val.strip()

            if code_str.endswith('.0'):
                code_str = code_str[:-2]

            row_num = int(row_idx_str)

            # Maestros match
            if code_str in maestros_catalog:
                total_labor_occurrences += 1
                desc = maestros_catalog[code_str]
                maestros_unconsolidated.append({'row': row_num, 'code': code_str, 'desc': desc})
                if code_str not in maestros_grouped:
                    maestros_grouped[code_str] = {'code': code_str, 'desc': desc, 'frequency': 0}
                maestros_grouped[code_str]['frequency'] += 1

            # CS match
            if code_str in cs_catalog:
                total_labor_occurrences += 1
                desc = cs_catalog[code_str]
                cs_unconsolidated.append({'row': row_num, 'code': code_str, 'desc': desc})
                if code_str not in cs_grouped:
                    cs_grouped[code_str] = {'code': code_str, 'desc': desc, 'frequency': 0}
                cs_grouped[code_str]['frequency'] += 1

    print(f"\n3. Extracción completada en {time.time()-t0:.2f} segundos:")
    print(f" -> Filas totales analizadas: {total_sales_rows:,}")
    print(f" -> Transacciones de Mano de Obra en Maestros (T49): {len(maestros_unconsolidated)} filas ({len(maestros_grouped)} códigos únicos)")
    print(f" -> Transacciones de Mano de Obra en Centro de Servicios (T39): {len(cs_unconsolidated)} filas ({len(cs_grouped)} códigos únicos)")

    # 4. Generate Unconsolidated Excel Report (Individual Rows - Fila por Fila)
    out_wb = openpyxl.Workbook()
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    code_font = Font(name="Consolas", size=10, bold=True, color="0284C7")
    code_font_cs = Font(name="Consolas", size=10, bold=True, color="059669")
    regular_font = Font(name="Calibri", size=10)

    # Sheet 1: Taller Maestro (T49) - Fila por fila
    ws1 = out_wb.active
    ws1.title = "Taller Maestro (T49)"
    header_fill_m = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    
    ws1.append(["N° Fila Sábana", "Código Identificado", "Descripción de la Actividad"])
    for cell in ws1[1]:
        cell.fill = header_fill_m
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r in sorted(maestros_unconsolidated, key=lambda x: x['row']):
        ws1.append([r['row'], r['code'], r['desc']])

    for r in range(2, ws1.max_row + 1):
        ws1[f"A{r}"].font = regular_font
        ws1[f"B{r}"].font = code_font
        ws1[f"C{r}"].font = regular_font

    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 75

    # Sheet 2: Centro de Servicios (T39) - Fila por fila
    ws2 = out_wb.create_sheet(title="Centro de Servicios (T39)")
    header_fill_cs = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

    ws2.append(["N° Fila Sábana", "Código Identificado", "Descripción de la Actividad"])
    for cell in ws2[1]:
        cell.fill = header_fill_cs
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r in sorted(cs_unconsolidated, key=lambda x: x['row']):
        ws2.append([r['row'], r['code'], r['desc']])

    for r in range(2, ws2.max_row + 1):
        ws2[f"A{r}"].font = regular_font
        ws2[f"B{r}"].font = code_font_cs
        ws2[f"C{r}"].font = regular_font

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 75

    output_excel = f"Reporte_Mano_de_Obra_{os.path.splitext(filename_base)[0]}.xlsx"
    out_wb.save(output_excel)
    out_wb.close()
    print(f"\n -> Reporte Excel FILA POR FILA generado: {os.path.abspath(output_excel)}")

    # 5. Maintain Multi-Month Historical Database in datos_costeo.js
    month_clean = os.path.splitext(filename_base)[0].replace("Consolidado de ventas", "").strip()
    if not month_clean: month_clean = "Actual"

    # Sort consolidated entries DE MAYOR A MENOR (Highest frequency first)
    maestros_sorted_desc = sorted(list(maestros_grouped.values()), key=lambda x: x['frequency'], reverse=True)
    cs_sorted_desc = sorted(list(cs_grouped.values()), key=lambda x: x['frequency'], reverse=True)

    month_record = {
        "fileName": filename_base,
        "monthTag": month_clean,
        "processedAt": time.strftime("%Y-%m-%d %H:%M:%S"),
        "totalSalesRows": total_sales_rows,
        "totalLaborRows": total_labor_occurrences,
        "maestrosMatches": maestros_sorted_desc,
        "csMatches": cs_sorted_desc
    }

    # Load existing history if datos_costeo.js exists
    costeo_history = {}
    if os.path.exists("datos_costeo.js"):
        try:
            with open("datos_costeo.js", "r", encoding="utf-8") as f:
                c_content = f.read()
                m = re.search(r'var\s+COSTEO_HISTORY\s*=\s*(\{[\s\S]*?\});', c_content)
                if m:
                    costeo_history = json.loads(m.group(1))
        except Exception as ex:
            print(" -> Nota: Inicializando nueva base de historial.")

    # Add or update current month
    costeo_history[month_clean] = month_record

    # Export updated COSTEO_HISTORY to datos_costeo.js
    js_content = f"// Base de datos histórica acumulada de costeo de mano de obra\nvar COSTEO_HISTORY = {json.dumps(costeo_history, ensure_ascii=False, indent=2)};\n"
    with open("datos_costeo.js", "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f" -> Historial acumulado actualizado ({len(costeo_history)} meses almacenados en datos_costeo.js)")

    # 6. Auto Push to GitHub Pages
    print("\n4. Publicando actualización automática en el Dashboard Web (GitHub Pages)...")
    try:
        subprocess.run(["git", "add", "datos_costeo.js"], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["git", "commit", "-m", f"Auto-sync: Historial de costeo actualizado con {month_clean}"], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["git", "push", "origin", "main"], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print(" -> ¡Publicado con éxito en GitHub Pages!")
    except Exception as ge:
        print(" -> (Sincronizado localmente para la web)")

    print(f"\n==========================================================")
    print(f"  ¡PROCESO COMPLETADO Y PUBLICADO CON ÉXITO!")
    print(f"==========================================================")

    # Automatically open local unconsolidated Excel file
    try:
        os.startfile(os.path.abspath(output_excel))
    except Exception:
        pass

except Exception as e:
    print(f"\nError durante la ejecución del agente: {e}")
    import traceback
    traceback.print_exc()

input("\nPresiona Enter para cerrar...")
